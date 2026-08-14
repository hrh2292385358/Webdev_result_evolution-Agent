"""ReportGenerator：生成完整的多 Sheet Excel 分析报告。"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ..config import REPORTS_DIR

# 颜色常量
C_HEADER   = "1B242F"
C_JADE     = "3FD6A4"
C_AMBER    = "E8A13C"
C_RED      = "E5604D"
C_SLATE    = "5B8BB0"
C_TXT      = "E6EDF3"
C_MUTED    = "8A98A8"
C_WHITE    = "FFFFFF"
C_DARK     = "0D1117"


def _header_font():  return Font(bold=True, color=C_TXT, name="Calibri")
def _header_fill():  return PatternFill("solid", fgColor=C_HEADER)
def _jade_font():    return Font(bold=True, color=C_JADE, name="Calibri")
def _red_font():     return Font(color=C_RED, name="Calibri")
def _amber_font():   return Font(color=C_AMBER, name="Calibri")

def _thin_border():
    s = Side(style="thin", color="26323F")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_header(ws, row: int, cols: list[str]):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row, c, val)
        cell.font  = _header_font()
        cell.fill  = _header_fill()
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        width = min_w
        for cell in col:
            try:
                width = max(width, min(len(str(cell.value or "")), max_w))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = width + 2


def _pct(v: Any) -> str:
    if v is None:
        return "—"
    try:
        return f"{float(v)*100:.1f}%"
    except Exception:
        return str(v)


def _fmt(v: Any) -> str:
    return "—" if v is None else str(v)


class ReportGenerator:
    def __init__(self, task_id: str, db):
        self.task_id = task_id
        self.db = db
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # 删除默认 Sheet

    def generate(self) -> tuple[Path, str]:
        from ..models import (
            Task, CheckResult, MetricResult,
            DifferenceRecord, Recommendation,
        )
        task = self.db.query(Task).filter(Task.id == self.task_id).first()
        name = (task.name if task else self.task_id).replace("/", "-")
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"Webdev_result_evolution_Report_{name}_{ts}.xlsx"
        path  = REPORTS_DIR / fname

        cr = self.db.query(CheckResult).filter(
            CheckResult.task_id == self.task_id,
            CheckResult.check_type == "overall",
        ).order_by(CheckResult.created_at.desc()).first()
        check_data = cr.details if cr else {}

        metrics = self.db.query(MetricResult).filter(
            MetricResult.task_id == self.task_id).all()
        overall_m = next((m.metrics for m in metrics if m.scope == "overall"), {}) or {}
        cat_m     = {m.scope_key: m.metrics for m in metrics if m.scope == "category"}
        dim_m     = {m.scope_key: m.metrics for m in metrics if m.scope == "dimension"}
        model_m   = {m.scope_key: m.metrics for m in metrics if m.scope == "model"}

        diffs = self.db.query(DifferenceRecord).filter(
            DifferenceRecord.task_id == self.task_id).all()
        recs  = self.db.query(Recommendation).filter(
            Recommendation.task_id == self.task_id).all()

        # ---- Sheets -------------------------------------------------------
        self._sheet_summary_and_metrics(task, ts, check_data, overall_m)
        self._sheet_category_metrics(cat_m)
        self._sheet_dimension_metrics(dim_m)
        self._sheet_model_metrics(model_m, overall_m.get("model_warnings", []))
        self._sheet_differences(diffs)
        self._sheet_skill_recs([r for r in recs if r.rec_type == "skill"])
        self._sheet_strategy_recs([r for r in recs if r.rec_type == "strategy"])
        self._sheet_weight_recs([r for r in recs if r.rec_type == "weight"])

        self.wb.save(path)
        return path, fname

    # ---- 00 任务摘要 + 整体指标（合并）----------------------------------------
    def _sheet_summary_and_metrics(self, task, ts, check, overall_m):
        ws = self.wb.create_sheet("00_任务摘要与整体指标")

        # 任务信息块
        ws.append(["任务信息", ""])
        _write_header(ws, ws.max_row, ["任务信息", ""])
        info_rows = [
            ("任务ID",      task.id if task else ""),
            ("任务名称",    task.name if task else ""),
            ("数据批次",    task.batch if task else ""),
            ("Rubric版本",  task.rubric_version if task else ""),
            ("Skill版本",   task.skill_version if task else ""),
            ("Judge模型",   task.judge_model if task else ""),
            ("评分范围",    task.score_range if task else ""),
            ("分析状态",    task.status if task else ""),
            ("报告生成时间",ts),
        ]
        for r in info_rows:
            ws.append(list(r))

        ws.append([])

        # 数据质量块
        dq = overall_m.get("data_quality") or {}
        ws.append(["数据质量统计", ""])
        _write_header(ws, ws.max_row, ["数据质量统计", ""])
        dq_rows = [
            ("成功匹配格数",   _fmt(dq.get("matched_cells"))),
            ("GT独有格",       _fmt(dq.get("gt_only_cells"))),
            ("Auto独有格",     _fmt(dq.get("auto_only_cells"))),
            ("Auto漏评格",     _fmt(dq.get("missing_auto_cells"))),
            ("非法评分格",     _fmt(dq.get("illegal_score_cells"))),
            ("豁免格",         _fmt(dq.get("exempt_cells"))),
        ]
        for r in dq_rows:
            ws.append(list(r))

        ws.append([])

        # 整体指标块
        ws.append(["整体指标", "值", "说明"])
        _write_header(ws, ws.max_row, ["整体指标", "值", "说明"])
        metric_rows = [
            ("有效评分格",   _fmt(overall_m.get("valid_cells")),     f"理论格：{_fmt(overall_m.get('theoretical_cells'))}"),
            ("无效评分格",   _fmt(overall_m.get("invalid_cells")),   "空值+豁免+非法"),
            ("有效格率",     _pct(overall_m.get("valid_rate")),      ""),
            ("覆盖率",       _pct(overall_m.get("coverage")),        "有效格/理论格"),
            ("精确一致率",   _pct(overall_m.get("exact_match")),     f"{overall_m.get('exact_match_num','—')}/{overall_m.get('valid_cells','—')}"),
            ("±1内一致率",   _pct(overall_m.get("within1")),         "abs(delta)≤1 的比例"),
            ("MAE",          _fmt(overall_m.get("mae")),             "平均绝对误差"),
            ("Bias",         _fmt(overall_m.get("bias")),            ">0偏乐观，<0偏严格"),
            ("Macro F1",     _pct(overall_m.get("macro_f1")),        ""),
            ("Weighted F1",  _pct(overall_m.get("weighted_f1")),     ""),
            ("严重误判数",   _fmt(overall_m.get("severe_errors")),   "GT=0↔Auto=2"),
            ("严重误判率",   _pct(overall_m.get("severe_error_rate")),""),
        ]
        for r in metric_rows:
            ws.append(list(r))

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 30

    # ---- 04 分类指标 --------------------------------------------------------
    def _sheet_category_metrics(self, cat_m):
        ws = self.wb.create_sheet("01_分类指标")
        cols = ["类别", "精确一致率", "±1一致率", "MAE", "Bias", "Macro F1", "有效格"]
        ws.append(cols); _write_header(ws, 1, cols)
        for cat, m in sorted(cat_m.items()):
            ws.append([cat, _pct(m.get("exact_match")), _pct(m.get("within1")),
                       _fmt(m.get("mae")), _fmt(m.get("bias")),
                       _pct(m.get("macro_f1")), _fmt(m.get("valid_cells"))])
        ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:G{len(cat_m)+1}"
        _auto_width(ws)

    # ---- 05 逐维度指标 ------------------------------------------------------
    def _sheet_dimension_metrics(self, dim_m):
        ws = self.wb.create_sheet("02_逐维度指标")
        cols = ["维度", "精确一致率", "±1一致率", "MAE", "Bias", "严重误判率", "有效格", "无效格", "有效格率"]
        ws.append(cols); _write_header(ws, 1, cols)
        for dim, m in sorted(dim_m.items()):
            ws.append([dim, _pct(m.get("exact_match")), _pct(m.get("within1")),
                       _fmt(m.get("mae")), _fmt(m.get("bias")),
                       _pct(m.get("severe_error_rate")), _fmt(m.get("valid_cells")),
                       _fmt(m.get("invalid_cells")), _pct(m.get("valid_rate"))])
        ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:I{len(dim_m)+1}"
        _auto_width(ws)

    # ---- 05b 模型指标 + 告警 ------------------------------------------------
    def _sheet_model_metrics(self, model_m: dict, warnings: list):
        ws = self.wb.create_sheet("03_模型指标")
        cols = ["候选模型", "精确一致率", "±1一致率", "MAE", "Bias",
                "有效格", "无效格", "有效格率", "空GT格", "空Auto格", "豁免格", "告警"]
        ws.append(cols); _write_header(ws, 1, cols)

        warn_map = {w["model"]: w for w in (warnings or [])}
        for model, m in sorted(model_m.items()):
            w = warn_map.get(model, {})
            issues = "；".join(w.get("suggestions", [])) if w else ""
            row = [
                model,
                _pct(m.get("exact_match")), _pct(m.get("within1")),
                _fmt(m.get("mae")), _fmt(m.get("bias")),
                _fmt(m.get("valid_cells")), _fmt(m.get("invalid_cells")),
                _pct(m.get("valid_rate")),
                _fmt(m.get("empty_gt_cells")), _fmt(m.get("empty_auto_cells")),
                _fmt(m.get("exempt_cells")),
                issues,
            ]
            ws.append(row)
            # 有效格率低于 50% 的行标红
            if (m.get("valid_rate") or 1.0) < 0.5:
                for cell in ws[ws.max_row]:
                    cell.font = _red_font()

        ws.freeze_panes = "A2"
        _auto_width(ws)

    # ---- 混淆矩阵（不生成 Sheet，但保留计算逻辑供其他地方调用）----------------
    def _compute_confusion(self, overall_m):
        return overall_m.get("confusion_matrix") or {}

    # ---- 01 差异样本明细 ----------------------------------------------------
    def _sheet_differences(self, diffs):
        ws = self.wb.create_sheet("01_差异样本明细")
        cols = ["数据ID","候选模型","维度","GT分","Auto分","差值","严重误判","归因","Auto原因"]
        ws.append(cols); _write_header(ws, 1, cols)
        for d in diffs:
            extra = d.extra or {}
            gt_val   = "-" if (extra.get("gt_is_exempt") or d.ground_truth_score is None) else d.ground_truth_score
            auto_val = "-" if (extra.get("auto_is_exempt") or not extra.get("has_auto", True) or d.auto_score is None) else d.auto_score
            ws.append([
                d.data_id, d.candidate_model, d.dimension_code,
                gt_val, auto_val, d.delta,
                "是" if d.is_severe_error else "",
                d.root_cause_category or "—",
                (d.auto_reason or "")[:200],
            ])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:I{len(diffs)+1}"
        _auto_width(ws)

    # ---- 08 Skill优化建议 ---------------------------------------------------
    def _sheet_skill_recs(self, recs):
        ws = self.wb.create_sheet("05_Skill优化建议")
        cols = ["维度","优先级","状态","问题描述","建议","LLM生成","LLM总结","优化后评分标准建议"]
        ws.append(cols); _write_header(ws, 1, cols)
        for r in recs:
            c = r.content or {}
            ws.append([
                r.target_dimension, r.priority, r.status,
                c.get("current_problem",""), c.get("suggestion",""),
                "是" if r.llm_generated else "否",
                c.get("llm_summary",""),
                c.get("suggested_rubric",""),
            ])
        ws.freeze_panes = "A2"; _auto_width(ws)

    # ---- 09 策略优化建议 ----------------------------------------------------
    def _sheet_strategy_recs(self, recs):
        ws = self.wb.create_sheet("06_策略优化建议")
        cols = ["目标问题","影响维度","当前策略","建议策略","时间成本","Token成本","优先级"]
        ws.append(cols); _write_header(ws, 1, cols)
        for r in recs:
            c = r.content or {}
            ws.append([
                c.get("target_problem",""),
                ", ".join(c.get("affected_dimensions",[])),
                c.get("current_strategy",""),
                c.get("suggested_strategy",""),
                c.get("time_cost",""), c.get("token_cost",""),
                r.priority,
            ])
        ws.freeze_panes = "A2"; _auto_width(ws)

    # ---- 10 权重调整建议 ----------------------------------------------------
    def _sheet_weight_recs(self, recs):
        ws = self.wb.create_sheet("07_权重调整建议")
        cols = ["方案","描述","类别","当前权重","建议权重","变化","是否推荐"]
        ws.append(cols); _write_header(ws, 1, cols)
        for r in recs:
            c = r.content or {}
            sw = c.get("suggested_weights", {})
            cw = c.get("current_weights", {})
            ch = c.get("changes", {})
            for cat in sorted(sw.keys()):
                ws.append([
                    c.get("scenario_name",""), c.get("description",""),
                    cat, cw.get(cat,"—"), sw.get(cat,"—"),
                    ch.get(cat,"—"),
                    "推荐" if c.get("recommended") else "",
                ])
        ws.freeze_panes = "A2"; _auto_width(ws)
