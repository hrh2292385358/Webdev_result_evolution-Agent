"""FileParser：读取Excel文件，检测Sheet、列、合并单元格等基础信息。"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter


class FileParseError(Exception):
    pass


class FileParser:
    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.wb: Optional[openpyxl.Workbook] = None

    def load(self) -> "FileParser":
        if not self.path.exists():
            raise FileParseError(f"文件不存在：{self.path}")
        try:
            self.wb = openpyxl.load_workbook(self.path, read_only=False, data_only=True)
        except Exception as e:
            raise FileParseError(f"无法打开文件：{e}")
        if not self.wb.sheetnames:
            raise FileParseError("文件中没有任何Sheet")
        return self

    def basic_info(self) -> dict:
        """返回文件基础信息：sheet列表、行列数等。"""
        info = {
            "file_name": self.path.name,
            "sheets": [],
        }
        for name in self.wb.sheetnames:
            ws = self.wb[name]
            info["sheets"].append({
                "name": name,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "merged_cells": len(ws.merged_cells.ranges),
            })
        return info

    def get_headers(self, sheet_name: Optional[str] = None, header_row: int = 1) -> list[str]:
        """读取指定Sheet第header_row行的表头。"""
        ws = self._ws(sheet_name)
        return [str(ws.cell(header_row, c).value or "") for c in range(1, (ws.max_column or 0) + 1)]

    def get_rows(self, sheet_name: Optional[str] = None, header_row: int = 1) -> list[dict[str, Any]]:
        """返回所有数据行（跳过全空行）。"""
        ws = self._ws(sheet_name)
        headers = self.get_headers(sheet_name, header_row)
        rows = []
        for r in range(header_row + 1, (ws.max_row or 0) + 1):
            vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
            if all(v is None or str(v).strip() == "" for v in vals):
                continue
            rows.append(dict(zip(headers, vals)))
        return rows

    def check_duplicate_headers(self, sheet_name: Optional[str] = None, header_row: int = 1) -> list[str]:
        """返回重复列名列表。"""
        headers = [h for h in self.get_headers(sheet_name, header_row) if h.strip()]
        seen, dupes = set(), []
        for h in headers:
            if h in seen:
                dupes.append(h)
            seen.add(h)
        return dupes

    def check_empty_score_cols(self, dim_codes: list[str], sheet_name: Optional[str] = None) -> list[str]:
        """返回评分列全部为空的维度代码列表。"""
        ws = self._ws(sheet_name)
        headers = self.get_headers(sheet_name)
        empty = []
        for code in dim_codes:
            col = self._find_col_by_code(headers, code)
            if col is None:
                continue
            vals = [ws.cell(r, col + 1).value for r in range(2, (ws.max_row or 0) + 1)]
            if all(v is None or str(v).strip() == "" for v in vals):
                empty.append(code)
        return empty

    def _find_col_by_code(self, headers: list[str], code: str) -> Optional[int]:
        """按维度代码找列索引（0-based）。"""
        pattern = re.compile(r'\b' + re.escape(code) + r'\b', re.IGNORECASE)
        for i, h in enumerate(headers):
            if pattern.search(h):
                return i
        return None

    def _ws(self, sheet_name: Optional[str] = None):
        if sheet_name and sheet_name in self.wb.sheetnames:
            return self.wb[sheet_name]
        return self.wb[self.wb.sheetnames[0]]

    def close(self):
        if self.wb:
            self.wb.close()
