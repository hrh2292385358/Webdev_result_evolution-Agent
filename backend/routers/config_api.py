"""Skills / Weights 配置读写 API。"""
from __future__ import annotations

import io
from pathlib import Path
from typing import Any

import yaml
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import JSONResponse

from ..models import ApiResponse

CONFIG_DIR = Path(__file__).resolve().parents[2] / "config"
DIMENSIONS_FILE = CONFIG_DIR / "dimensions.yaml"
WEIGHTS_FILE = CONFIG_DIR / "weights.yaml"

router = APIRouter()


def _read_yaml(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        raw = path.read_text(encoding="utf-8")
        # 剥掉顶部 Python docstring（历史遗留）
        if raw.lstrip().startswith('"""') or raw.lstrip().startswith("'''"):
            raw = raw[raw.index("\n") + 1:]
        return yaml.safe_load(raw) or {}
    except Exception:
        return {}


def _write_yaml(path: Path, data: dict):
    path.write_text(yaml.dump(data, allow_unicode=True, default_flow_style=False,
                              sort_keys=False), encoding="utf-8")


def _parse_md_skills(text: str) -> dict:
    """从 Markdown 维度说明文件解析完整 dimensions 列表。
    支持格式：章节标题含类别，子节标题含维度代码，属性表格含字段。
    """
    import re
    category_map = [
        (re.compile(r'Gateway|门槛层', re.IGNORECASE), 'Gateway'),
        (re.compile(r'Functionality|功能层', re.IGNORECASE), 'Functionality'),
        (re.compile(r'Interactivity|交互层', re.IGNORECASE), 'Interactivity'),
        (re.compile(r'Aesthetics|美观层', re.IGNORECASE), 'Aesthetics'),
        (re.compile(r'Content|内容层', re.IGNORECASE), 'Content'),
        (re.compile(r'DataPersistence|数据层|数据持久', re.IGNORECASE), 'DataPersistence'),
    ]
    scale_map = {
        re.compile(r'0/1/2|三档|三分'): '0/1/2',
        re.compile(r'0/1\b|二分|pass.*fail|二值'): '0/1',
    }
    code_re = re.compile(r'\b(G[1-4]|F[1-4]|DP[1-4]|I[1-4]|A[1-4]|C[1-2])\b')

    dims: dict[str, dict] = {}
    current_cat = None
    current_code = None
    rubric_lines: list[str] = []
    in_rubric = False

    def _flush_rubric():
        if current_code and rubric_lines:
            dims[current_code]['rubric_points'] = '\n'.join(rubric_lines).strip()
        rubric_lines.clear()

    lines = text.splitlines()
    for i, line in enumerate(lines):
        stripped = line.strip()

        # 二级标题 → 切换类别
        if stripped.startswith('## '):
            _flush_rubric()
            in_rubric = False
            for pat, cat in category_map:
                if pat.search(stripped):
                    current_cat = cat
                    break

        # 三级标题 → 切换维度
        elif stripped.startswith('### '):
            _flush_rubric()
            in_rubric = False
            m = code_re.search(stripped)
            if m and current_cat:
                current_code = m.group(1).upper()
                name = re.sub(r'^###\s*', '', stripped)
                name = re.sub(r'\b' + re.escape(current_code) + r'\b\s*[·\-\s]*', '', name).strip()
                dims[current_code] = {
                    'code': current_code,
                    'category': current_cat,
                    'name': name,
                    'layer': 'gateway' if current_cat == 'Gateway' else 'scoring',
                    'scale': '0/1' if current_cat == 'Gateway' else '0/1/2',
                    'automatable': False,
                    'eval_type': '',
                    'exemption': '',
                    'rubric_points': '',
                }

        elif current_code:
            d = dims[current_code]

            # 评分标准段开始（优先判断，避免被属性表格拦截）
            if stripped.startswith('**') and ('评分标准' in stripped or '判定标准' in stripped):
                in_rubric = True
                rubric_lines.clear()
                continue

            # 豁免规则段落
            if '豁免规则' in stripped or '豁免内容' in stripped:
                in_rubric = False
                exemption_parts = []
                j = i + 1
                while j < len(lines):
                    nl = lines[j].strip()
                    if nl.startswith('#') or (nl.startswith('**') and ('评分' in nl or '校准' in nl or '维度边界' in nl)):
                        break
                    if nl.startswith('---'):
                        break
                    if nl and not nl.startswith('|---'):
                        exemption_parts.append(nl)
                    j += 1
                d['exemption'] = ' '.join(exemption_parts).strip()
                continue

            if in_rubric:
                if stripped.startswith('---'):
                    _flush_rubric()
                    in_rubric = False
                elif stripped.startswith('**') and ('豁免规则' in stripped or '豁免内容' in stripped):
                    _flush_rubric()
                    in_rubric = False
                elif stripped.startswith('|---'):
                    pass  # 表格分隔行跳过
                elif stripped:
                    rubric_lines.append(stripped)
                continue

            # 属性表格行（仅在非 rubric 模式下解析）：| 属性 | 值 |
            if '|' in stripped and not stripped.startswith('|---') and not in_rubric:
                cells = [c.strip() for c in stripped.split('|') if c.strip()]
                if len(cells) >= 2:
                    key, val = cells[0], cells[1]
                    if key in ('量表', 'scale', '评分范围'):
                        for pat, sv in scale_map.items():
                            if pat.search(val):
                                d['scale'] = sv; break
                    elif key in ('可自动化', 'automatable'):
                        d['automatable'] = val in ('是', 'true', 'True', 'yes', '1')
                    elif key in ('评估类型', 'eval_type'):
                        d['eval_type'] = val

    _flush_rubric()

    # 按标准顺序排列
    order = ['G1','G2','G3','G4','F1','F2','F3','F4','DP1','DP2','DP3','DP4',
             'I1','I2','I3','I4','A1','A2','A3','A4','C1','C2']
    sorted_dims = [dims[c] for c in order if c in dims]
    sorted_dims += [v for k, v in dims.items() if k not in order]
    return {'dimensions': sorted_dims}


def _parse_excel_skills(content: bytes) -> dict:
    """从 Excel 读取 skills，期望列：code, category, name, layer, scale"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    dims = []
    for row in rows[1:]:
        if not any(row):
            continue
        d = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        code = str(d.get("code") or "").strip()
        if not code:
            continue
        dims.append({
            "code": code,
            "category": str(d.get("category") or "").strip(),
            "name": str(d.get("name") or "").strip(),
            "layer": str(d.get("layer") or "scoring").strip(),
            "scale": str(d.get("scale") or "0/1/2").strip(),
        })
    wb.close()
    return {"dimensions": dims}


def _parse_excel_weights(content: bytes) -> dict:
    """从 Excel 读取权重，期望列：category, weight"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    weights = {}
    for row in rows[1:]:
        if not any(row):
            continue
        d = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        cat = str(d.get("category") or "").strip()
        try:
            w = float(d.get("weight") or 0)
        except (TypeError, ValueError):
            continue
        if cat:
            weights[cat] = w
    wb.close()
    return {"category_weights": weights}


# ---------------------------------------------------------------------------
# Skills (dimensions.yaml)
# ---------------------------------------------------------------------------

@router.get("/config/skills", response_model=ApiResponse)
def get_skills():
    data = _read_yaml(DIMENSIONS_FILE)
    return ApiResponse(success=True, data=data)


@router.post("/config/skills", response_model=ApiResponse)
def save_skills(payload: dict):
    try:
        _write_yaml(DIMENSIONS_FILE, payload)
        return ApiResponse(success=True, data={"saved": True})
    except Exception as e:
        return ApiResponse(success=False, error=str(e))


@router.post("/config/skills/upload", response_model=ApiResponse)
async def upload_skills(file: UploadFile = File(...)):
    content = await file.read()
    fname = (file.filename or "").lower()
    try:
        if fname.endswith((".xlsx", ".xls")):
            data = _parse_excel_skills(content)
        elif fname.endswith(".md"):
            data = _parse_md_skills(content.decode("utf-8"))
        else:
            # .yaml / .yml / .txt
            data = yaml.safe_load(content.decode("utf-8")) or {}
        if not data.get("dimensions"):
            return ApiResponse(success=False, error="文件中未找到 dimensions 数据，请确保包含 dimensions: 列表")
        _write_yaml(DIMENSIONS_FILE, data)
        return ApiResponse(success=True, data=data)
    except Exception as e:
        return ApiResponse(success=False, error=str(e))


# ---------------------------------------------------------------------------
# Weights (weights.yaml)
# ---------------------------------------------------------------------------

@router.get("/config/weights", response_model=ApiResponse)
def get_weights():
    data = _read_yaml(WEIGHTS_FILE)
    return ApiResponse(success=True, data=data)


@router.post("/config/weights", response_model=ApiResponse)
def save_weights(payload: dict):
    try:
        _write_yaml(WEIGHTS_FILE, payload)
        return ApiResponse(success=True, data={"saved": True})
    except Exception as e:
        return ApiResponse(success=False, error=str(e))


@router.post("/config/weights/upload", response_model=ApiResponse)
async def upload_weights(file: UploadFile = File(...)):
    content = await file.read()
    fname = (file.filename or "").lower()
    try:
        if fname.endswith((".xlsx", ".xls")):
            data = _parse_excel_weights(content)
        else:
            # .yaml / .yml / .txt / .md — 都当 YAML 纯文本解析
            data = yaml.safe_load(content.decode("utf-8")) or {}
        if not data.get("category_weights"):
            return ApiResponse(success=False, error="文件中未找到 category_weights 数据，请确保包含 category_weights: 字段")
        _write_yaml(WEIGHTS_FILE, data)
        return ApiResponse(success=True, data=data)
    except Exception as e:
        return ApiResponse(success=False, error=str(e))


# ---------------------------------------------------------------------------
# 版本元信息（rubric/skill/weight version）
# ---------------------------------------------------------------------------

_META_FILE = CONFIG_DIR / "meta.yaml"


def _read_meta() -> dict:
    return _read_yaml(_META_FILE)


@router.get("/config/meta", response_model=ApiResponse)
def get_meta():
    return ApiResponse(success=True, data=_read_meta())


@router.post("/config/meta", response_model=ApiResponse)
def save_meta(payload: dict):
    try:
        _write_yaml(_META_FILE, payload)
        return ApiResponse(success=True, data={"saved": True})
    except Exception as e:
        return ApiResponse(success=False, error=str(e))
